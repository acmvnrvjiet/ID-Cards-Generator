import subprocess
import sys
# os package to run operating system commands through python(pre installed)
import os


def install(package):
    subprocess.check_call([sys.executable, "-m", "pip", "install", package])


install('pyqrcode')
install('openpyxl')
install('Pillow')
install('pypng')

import pyqrcode
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont
from tkinter.filedialog import askopenfile
# openpyxl package for accessing excel sheets

# pyqrcode package to generate qr codes


try:
    file = askopenfile(title='Select the Workbook', mode='r', filetypes=[
        ('Microsoft Excel', '.xlsx .xlsx .xlsm .xltx .xltm')])
except:
    {}

if file is not None:
    dirpath = os.path.dirname(file.name)
    wbname = os.path.basename(file.name)
else:
    sys.exit()

# filepath variable contains the path of the members info excel sheet
# r'' makes the string in raw format so that special characters like \n,\t will nt b taken into consideration
filepath = file.name

# dirpath variable contains the address of location where we are going to save th qr codes and txt file
# folder is named as QR_Codes
dirpath = dirpath + '/' + 'QR_Codes'

allIdsPath = dirpath + '/All_IDs'

# os.mkdir(filepath) is used to create a folder/dir where we are saving the qr codes
# it's written in try-except block so that if the folder is already present then exception is generated
# but will not cause any problem to the execution of the program
try:
    os.mkdir(dirpath)
    os.mkdir(allIdsPath)
except:
    {}

# wb variable stores the workbook of the excel file opened
# load_workbook opens the workbook
wb = load_workbook(filepath)

print(end='\n\n')
print(wbname, end='\n\n')

# for every worksheet(ws) in the opened workbook(wb):
for ws in wb:

    print(ws.title)

    # wspath stores the path of another inside folder with the name of the worksheet
    # .strip() is used to remove extra spaces at the ends of a string
    wspath = dirpath + '/' + ws.title.strip()

    #  it creates another folder/dir inside QR_Codes folder with the name of the worksheet ws
    try:
        os.mkdir(wspath)
    except:
        {}

    # for every row in the worksheet which starts from 3 to 60
    # as the info of members starts from row no 3
    # at max no. of students could be 60
    for r in range(3, 60):
        # cell variable selects the cell which is placed at location row=r n column=1
        cell = ws.cell(row=r, column=1)

        # if the first cell(sl_no cell) is empty the loop will break
        # cell.value gives the data inside the cell
        if cell.value is None:
            break

        # rno,name,pno,mail are variables which store the respective data of a member of row=r
        # .strip() is used to remove extra spaces at d ends of a string
        rno = str(ws.cell(row=r, column=2).value).strip()
        name = ws.cell(row=r, column=3).value.strip().replace('.', '')
        fname = ws.cell(row=r, column=4).value.strip().replace('.', '')
        lname = ws.cell(row=r, column=5).value.strip().replace('.', '')
        pno = str(int(ws.cell(row=r, column=6).value)).strip()
        mail = ws.cell(row=r, column=7).value.strip()

        # Use a pattern to create the required ID
        # Write your code for the ID pattern here
        acm_id = 'ID'+str(r-2)

        # eachmemberpath variable stores the path address where we are going to store the qrcode
        eachmemberpath = wspath + '/' + acm_id

        # it creates a folder with name as acm_id inside the sheet name folder
        try:
            os.mkdir(eachmemberpath)
        except:
            {}

        # qr_info variable stores information to be stored in the qr code
        # roll,name,pno,mail are concatenated with ','
        qr_info = str(rno) + ',' + str(fname) + ',' + \
                  str(lname) + ',' + str(pno) + ',' + str(mail)

        # qr_code variable stores the qr code image
        # pyqrcode.create(info) function creates qr code wid the given info
        qr_code = pyqrcode.create(qr_info)

        qrpath = eachmemberpath + '/' + acm_id + '_qr.png'
        # idpath stores the path where the id is to be stored with the name as acm_id and extension .png
        idpath = eachmemberpath + '/' + acm_id + '_id.png'

        # this line saves the qr_code image in png format at imagepath location
        qr_code.png(qrpath, scale=7)

        qr_code = Image.open(qrpath)
        qr_code = qr_code.resize((315, 315))  # Change the dimension values according to your requirements
        qr_code = qr_code.crop((24, 24, 291, 291))
        qr_code.save(qrpath)

        template = Image.open('sample.jpg')
        qr_code = Image.open(qrpath)
        template.paste(qr_code, (950, 480))
        # template.show()

        draw = ImageDraw.Draw(template)
        id_font = ImageFont.truetype('fonts/Blogger_Sans.ttf', 50)
        name_font = ImageFont.truetype('fonts/ArmWrestler.ttf', 60)
        pno_font = ImageFont.truetype('fonts/Amaranth-Bold.ttf', 50)

        # Here you can replace the dimension,color values according to your template image
        draw.text((133, 335), acm_id, fill=(0, 0, 0, 255), font=id_font)
        draw.text((85, 512), name.upper(), fill=(0, 0, 0, 255), font=name_font)
        draw.text((128, 585), 'PH.NO:' + pno, fill=(0, 0, 0, 255), font=pno_font)
        # template.show()
        template.save(idpath)
        template.save(allIdsPath + '/' + acm_id + '.png')

        # fp variable stores the address of a txt file containing the qr_info which is to be saved beside the qr code image
        fp = eachmemberpath + '/' + acm_id + '.txt'

        # file_acm stores a file opened at fp location in write mode
        # if the file already exists it overwrites it
        file_acm = open(fp, 'w')

        # lines variable contains the info to be stored in the file_acm file
        lines = [rno, '\n', fname, '\n', lname, '\n', pno, '\n', mail]

        # file.writelines(list of lines) function writes the given lines on the file_acm
        file_acm.writelines(lines)

        # file_acm is closed
        file_acm.close()

        # it continues with the next row

        # it continues with the next sheet in wb
        print(name, 'Done.')
    # print(ws.title, 'Sheet', 'Done.')
    print()
# after the job is done ... a Done statement is printed in the logs...!!!
print(wbname, 'WorkBook Done.', end='\n\n')
